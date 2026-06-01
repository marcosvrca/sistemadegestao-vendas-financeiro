import re
import unicodedata
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from schemas import FORMAS_PAGAMENTO_VENDA, FORMA_PAGAMENTO_AV

CAMPOS_OBRIGATORIOS = ["data", "valor", "forma_pag", "quantidade", "desconto", "produto"]
CAMPOS_OPCIONAIS = ["id", "cliente", "troco", "valor_recebido", "observacao"]

MAPEAMENTO_COLUNAS: dict[str, list[str]] = {
    "id": ["id", "codigo", "cod", "numero", "n"],
    "data": ["data", "date", "dt", "data_venda"],
    "valor": ["valor", "total", "valor_total", "venda", "valor_venda"],
    "forma_pag": [
        "forma_pag",
        "forma pag",
        "forma_pagamento",
        "forma pagamento",
        "pagamento",
        "forma de pagamento",
        "form_pag",
    ],
    "quantidade": ["quantidade", "qtd", "qtde", "qty"],
    "desconto": ["desconto", "desc", "valor_desconto"],
    "produto": ["produto", "item", "descricao", "descrição", "nome_produto", "artigo"],
    "cliente": ["cliente", "nome_cliente", "comprador"],
    "troco": ["troco", "valor_troco", "troco_devolvido"],
    "valor_recebido": [
        "valor_recebido",
        "valor recebido",
        "recebido",
        "valor_pago",
        "valor pago",
        "passou",
        "cliente_passou",
        "cliente passou",
        "valor_entregue",
        "valor entregue",
    ],
    "observacao": [
        "observacao",
        "observação",
        "obs",
        "anotacao",
        "anotação",
        "anotacoes",
        "anotações",
        "notas",
        "nota",
        "comentario",
        "comentário",
    ],
}

MAPA_FORMA_PAGAMENTO = {
    "dinheiro": "Dinheiro",
    "pix": "Pix",
    "cartao debito": "Cartão Débito",
    "cartao de debito": "Cartão Débito",
    "debito": "Cartão Débito",
    "deb": "Cartão Débito",
    "cartao credito": "Cartão Crédito",
    "cartao de credito": "Cartão Crédito",
    "credito": "Cartão Crédito",
    "cred": "Cartão Crédito",
    "outro": "Outro",
    "outros": "Outro",
    "av": FORMA_PAGAMENTO_AV,
    "a vencer": FORMA_PAGAMENTO_AV,
    "pagar depois": FORMA_PAGAMENTO_AV,
}


def normalizar_texto(valor: str) -> str:
    texto = unicodedata.normalize("NFKD", valor.strip().lower())
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto)
    return texto


def normalizar_coluna(valor: str) -> str:
    return normalizar_texto(valor).replace(" ", "_")


def detectar_colunas(headers: list[Any]) -> dict[str, int]:
    mapeamento: dict[str, int] = {}
    headers_norm = [normalizar_coluna(str(h)) if h is not None else "" for h in headers]

    for campo, aliases in MAPEAMENTO_COLUNAS.items():
        aliases_norm = [normalizar_coluna(a) for a in aliases]
        for idx, header in enumerate(headers_norm):
            if header in aliases_norm:
                mapeamento[campo] = idx
                break

    return mapeamento


def validar_colunas(mapeamento: dict[str, int]) -> list[str]:
    faltando = []
    for campo in CAMPOS_OBRIGATORIOS:
        if campo not in mapeamento:
            faltando.append(campo)
    return faltando


def parse_numero(valor: Any) -> float | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, bool):
        return float(valor)
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto or texto.lower() in {"-", "—", "n/a", "na"}:
        return None
    texto = texto.replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def parse_inteiro(valor: Any) -> int | None:
    numero = parse_numero(valor)
    if numero is None:
        return None
    return int(numero)


def parse_data(valor: Any) -> datetime | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor
    if hasattr(valor, "year"):
        return datetime(valor.year, valor.month, valor.day)

    texto = str(valor).strip()
    formatos = [
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y",
    ]
    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            continue
    return None


def normalizar_forma_pagamento(valor: Any) -> str | None:
    if valor is None or valor == "":
        return None
    chave = normalizar_texto(str(valor))
    if chave in MAPA_FORMA_PAGAMENTO:
        return MAPA_FORMA_PAGAMENTO[chave]
    for forma in FORMAS_PAGAMENTO_VENDA:
        if normalizar_texto(forma) == chave:
            return forma
    for alias, forma in MAPA_FORMA_PAGAMENTO.items():
        if alias in chave or chave in alias:
            return forma
    return None


def obter_celula(row: tuple, indice: int | None) -> Any:
    if indice is None or indice >= len(row):
        return None
    return row[indice]


def linha_vazia(row: tuple) -> bool:
    return all(c is None or str(c).strip() == "" for c in row)


def parsear_planilha(conteudo: bytes) -> tuple[list[dict], list[dict], dict[str, int]]:
    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Planilha vazia")

    header_idx = None
    mapeamento: dict[str, int] = {}

    for idx, row in enumerate(rows[:20]):
        if linha_vazia(row):
            continue
        candidato = detectar_colunas(list(row))
        if len(candidato) >= 4:
            header_idx = idx
            mapeamento = candidato
            break

    if header_idx is None:
        raise ValueError(
            "Não foi possível identificar o cabeçalho. "
            f"Colunas obrigatórias: {', '.join(CAMPOS_OBRIGATORIOS)}"
        )

    faltando = validar_colunas(mapeamento)
    if faltando:
        raise ValueError(
            f"Colunas obrigatórias não encontradas: {', '.join(faltando)}. "
            f"Encontradas: {', '.join(mapeamento.keys())}"
        )

    vendas: list[dict] = []
    erros: list[dict] = []

    for linha_num, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if linha_vazia(row):
            continue

        produto = obter_celula(row, mapeamento.get("produto"))
        if produto is None or str(produto).strip() == "":
            continue

        produto_str = str(produto).strip()
        if produto_str.lower() in {"produto", "total", "observação", "observacao"}:
            continue

        try:
            data = parse_data(obter_celula(row, mapeamento.get("data")))
            valor = parse_numero(obter_celula(row, mapeamento.get("valor")))
            quantidade = parse_inteiro(obter_celula(row, mapeamento.get("quantidade")))
            desconto_raw = obter_celula(row, mapeamento.get("desconto"))
            desconto = parse_numero(desconto_raw)
            forma_pag = normalizar_forma_pagamento(
                obter_celula(row, mapeamento.get("forma_pag"))
            )
            troco = parse_numero(obter_celula(row, mapeamento.get("troco")))
            valor_recebido = parse_numero(
                obter_celula(row, mapeamento.get("valor_recebido"))
            )
            venda_id = parse_inteiro(obter_celula(row, mapeamento.get("id")))
            cliente_raw = obter_celula(row, mapeamento.get("cliente"))
            observacao_raw = obter_celula(row, mapeamento.get("observacao"))

            if data is None:
                raise ValueError("Data inválida ou vazia")
            if valor is None:
                raise ValueError("Valor inválido ou vazio")
            if quantidade is None or quantidade < 1:
                raise ValueError("Quantidade inválida (mínimo 1)")
            if desconto is None:
                desconto = 0.0
            if forma_pag is None:
                raise ValueError("Forma de pagamento inválida ou vazia")

            if valor_recebido is not None and troco is not None:
                esperado = round(valor_recebido - troco, 2)
                if abs(esperado - valor) > 0.05:
                    observacao_extra = (
                        f"Troco informado: recebido R${valor_recebido:.2f}, "
                        f"troco R${troco:.2f} (esperado valor R${esperado:.2f})"
                    )
                else:
                    observacao_extra = None
            elif valor_recebido is not None and troco is None:
                troco = round(max(valor_recebido - valor, 0), 2)
                observacao_extra = None
            elif troco is not None and valor_recebido is None:
                valor_recebido = round(valor + troco, 2)
                observacao_extra = None
            else:
                observacao_extra = None

            valor_unitario = round((valor + desconto) / quantidade, 2)

            observacoes = []
            if observacao_raw and str(observacao_raw).strip():
                observacoes.append(str(observacao_raw).strip())
            if observacao_extra:
                observacoes.append(observacao_extra)

            vendas.append(
                {
                    "id": venda_id,
                    "data": data,
                    "produto": produto_str[:200],
                    "cliente": str(cliente_raw).strip()[:200] if cliente_raw else "Cliente avulso",
                    "quantidade": quantidade,
                    "valor_unitario": valor_unitario,
                    "desconto": desconto,
                    "valor": round(valor, 2),
                    "forma_pagamento": forma_pag,
                    "troco": troco,
                    "valor_recebido": valor_recebido,
                    "observacao": " | ".join(observacoes) if observacoes else None,
                    "linha": linha_num,
                }
            )
        except ValueError as exc:
            erros.append({"linha": linha_num, "mensagem": str(exc), "produto": produto_str})

    wb.close()
    return vendas, erros, mapeamento
